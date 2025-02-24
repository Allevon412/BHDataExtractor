"""
To utilize this example please install requests. The rest of the dependencies are part of the Python 3 standard
library.

# pip install --upgrade requests

Note: this script was written for Python 3.6.X or greater.

Insert your BHE API creds in the BHE constants and change the PRINT constants to print desired data.
"""

import hmac
import hashlib
import base64
import requests
import datetime
import json
import argparse
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

from typing import Optional

PRINT_PRINCIPALS = False
PRINT_ATTACK_PATH_TIMELINE_DATA = False
PRINT_POSTURE_DATA = False

DATA_START = "1970-01-01T00:00:00.000Z"
DATA_END = datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'  # Now


class Credentials(object):
    def __init__(self, token_id: str, token_key: str) -> None:
        self.token_id = token_id
        self.token_key = token_key


class APIVersion(object):
    def __init__(self, api_version: str, server_version: str) -> None:
        self.api_version = api_version
        self.server_version = server_version


class Domain(object):
    def __init__(self, name: str, id: str, collected: bool, domain_type: str) -> None:  # impact_value: int) -> None:
        self.name = name
        self.id = id
        self.type = domain_type
        self.collected = collected
        # self.impact_value = impact_value


class Client(object):
    def __init__(self, scheme: str, host: str, port: int, credentials: Credentials) -> None:
        self._scheme = scheme
        self._host = host
        self._port = port
        self._credentials = credentials

    def _format_url(self, uri: str) -> str:
        formatted_uri = uri
        if uri.startswith("/"):
            formatted_uri = formatted_uri[1:]

        return f"{self._scheme}://{self._host}:{self._port}/{formatted_uri}"

    def _request(self, method: str, uri: str, body: Optional[bytes] = None) -> requests.Response:
        # Digester is initialized with HMAC-SHA-256 using the token key as the HMAC digest key.
        digester = hmac.new(self._credentials.token_key.encode(), None, hashlib.sha256)

        # OperationKey is the first HMAC digest link in the signature chain. This prevents replay attacks that seek to
        # modify the request method or URI. It is composed of concatenating the request method and the request URI with
        # no delimiter and computing the HMAC digest using the token key as the digest secret.
        #
        # Example: GET /api/v1/test/resource HTTP/1.1
        # Signature Component: GET/api/v1/test/resource
        digester.update(f"{method}{uri}".encode())

        # Update the digester for further chaining
        digester = hmac.new(digester.digest(), None, hashlib.sha256)

        # DateKey is the next HMAC digest link in the signature chain. This encodes the RFC3339 formatted datetime
        # value as part of the signature to the hour to prevent replay attacks that are older than max two hours. This
        # value is added to the signature chain by cutting off all values from the RFC3339 formatted datetime from the
        # hours value forward:
        #
        # Example: 2020-12-01T23:59:60Z
        # Signature Component: 2020-12-01T23
        datetime_formatted = datetime.datetime.now().astimezone().isoformat("T")
        digester.update(datetime_formatted[:13].encode())

        # Update the digester for further chaining
        digester = hmac.new(digester.digest(), None, hashlib.sha256)

        # Body signing is the last HMAC digest link in the signature chain. This encodes the request body as part of
        # the signature to prevent replay attacks that seek to modify the payload of a signed request. In the case
        # where there is no body content the HMAC digest is computed anyway, simply with no values written to the
        # digester.
        if body is not None:
            digester.update(body)

        # Perform the request with the signed and expected headers
        return requests.request(
            method=method,
            url=self._format_url(uri),
            headers={
                "User-Agent": "bhe-python-sdk 0001",
                "Authorization": f"bhesignature {self._credentials.token_id}",
                "RequestDate": datetime_formatted,
                "Signature": base64.b64encode(digester.digest()),
                "Content-Type": "application/json",
            },
            data=body,
        )

    def get_version(self) -> APIVersion:
        response = self._request("GET", "/api/version")
        payload = response.json()

        return APIVersion(api_version=payload["data"]["API"]["current_version"],
                          server_version=payload["data"]["server_version"])

    def get_domains(self) -> list[Domain]:
        response = self._request('GET', '/api/v2/available-domains')
        payload = response.json()['data']

        domains = list()
        for domain in payload:
            domains.append(
                Domain(domain["name"], domain["id"], domain["collected"], domain["type"]))  # domain["impactValue"]

        return domains

    def run_cypher(self, query, include_properties=False) -> requests.Response:
        """ Runs a Cypher query and returns the results

        Parameters:
        query (string): The Cypher query to run
        include_properties (bool): Should all properties of result nodes/edges be returned

        Returns:
        string: JSON result

        """

        data = {
            "include_properties": include_properties,
            "query": query
        }
        body = json.dumps(data).encode('utf8')
        response = self._request("POST", "/api/v2/graphs/cypher", body)
        return response.json()

    def get_ad_entity(self, entity_id: str) -> requests.Response:
        """ Gets an AD entity by its ID

        Parameters:
        entity_id (string): The ID of the entity to get

        Returns:
        string: JSON result

        """
        response = self._request("GET", f"/api/v2/base/{entity_id}")
        return json.loads(response.text)['data']

    def get_computer(self, computer_id: str) -> requests.Response:
        """ Gets a computer by its ID

        Parameters:
        computer_id (string): The ID of the computer to get

        Returns:
        string: JSON result

        """
        response = self._request("GET", f"/api/v2/computers/{computer_id}")
        return json.loads(response.text)['data']


class Parser:
    def __init__(self, cypher_name, workbook, client, wbname):
        self.name = cypher_name
        self.wb = workbook
        self.client = client
        self.wbname = wbname

    def parse(self, cypher_result):

        try:
            status = cypher_result['http_status']
            if status == 404:
                print(f"[*] [{self.name}]: no nodes detected")
                return
            if status == 400:
                print(f"[*] [{self.name}]: bad request check cypher query")
                return
            if status == 500:
                print(f"[*] [{self.name}]: internal server error")
                return
            if status == 429:
                print(f"[*] [{self.name}]: rate limit exceeded")
                return

        except KeyError:
            pass

        if self.name == "Kerberoastable users":

            # Get nodes from Cypher result
            self.standard_write(cypher_result)

        if self.name == "SPNs with Special Privileges":
            self.standard_write(cypher_result)

        elif self.name == "AS-REP Roastable users":
            self.standard_write(cypher_result)

        elif self.name == "High Value principals AS-REP Roastable users":
            self.standard_write(cypher_result)

        elif self.name == "Disabled Tier Zero High Value principals":
            self.standard_write(cypher_result)

        elif self.name == "Tier Zero High Value users with non-expiring passwords":
            self.standard_write(cypher_result)

        elif self.name == "Enabled Users w Passwrd not rotated in 365 days":
            self.standard_write(cypher_result)

        elif self.name == "Users which do not require password to authenticate":
            self.standard_write(cypher_result)

        elif self.name == "High Value principals which do not require password":
            self.standard_write(cypher_result)
        elif self.name == "Enabled Tier Zero High Value principals inactive for 60 days":
            self.standard_write(cypher_result)
        elif self.name == "Users with \"pass\" in their description":
            if len(self.name) > 31:
                sheet = self.wb.create_sheet(title=self.name[:30])
            else:
                sheet = self.wb.create_sheet(title=self.name)

            nodes = cypher_result['data']['nodes']
            if len(nodes.items()) > 1000: # if the response has too many hosts it'll run for days trying to get the host info.
                self.standard_write(cypher_result)
                return
            max_threads = 20
            data_list = list()
            if nodes:
                with ThreadPoolExecutor(max_threads) as executor:
                    futures = {executor.submit(self.client.get_ad_entity, node_data['objectId']): node_data for
                               node_id, node_data in nodes.items()}
                    for future in as_completed(futures):
                        data_list.append(future.result())

                count = 1
                sheet.cell(row=count, column=1).value = self.name
                count += 1
                for data in data_list:
                    sheet.cell(row=count, column=1).value = (f"[User Object]: {data['props']['name']} - [Description]: {data['props']['description']}\n")
                    count += 1

        elif self.name == "Computers with Unconstrained Delegation":
            # remove all the domain controllers from our data set. I tried to filter this out using cypher but nothing I was trying was working.
            # in stead I use the tier zero flag to filter out the domain controllers.
            # WARNING: this flag can be set manually by a user in the BH database. By default though domain controllers are set with this flag.
            data = cypher_result['data']['nodes']
            data_without_domain_controllers = dict()
            for node_id, node_data in data.items():
                if node_data["isTierZero"]:
                    continue
                else:
                    data_without_domain_controllers[node_id] = node_data
            cypher_result['data']['nodes'] = data_without_domain_controllers
            self.standard_write(cypher_result)

        elif self.name == "Principals with passwords stored using reversible encryption":
            self.standard_write(cypher_result)

        elif self.name == "ADCS ESC1" or self.name == "ADCS ESC2":
            if len(self.name) > 31:
                sheet = self.wb.create_sheet(title=self.name[:30])
            else:
                sheet = self.wb.create_sheet(title=self.name)

            sheet.column_dimensions['A'].width = 50
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 100
            parsed_data = self.parse_adcs_edges(cypher_result)
            count = 1
            sheet.cell(row=count, column=1).value = self.name
            count += 1
            for node_id, node_data in parsed_data.items():
                sheet.cell(row=count, column=1).value = (node_data['label'])
                sheet.cell(row=count, column=2).value = (node_data['kind'])
                try:
                    sheet.cell(row=count, column=3).value = (node_data['edges'])
                except KeyError:
                    pass
                count += 1


    def standard_write(self, data):
        if len(self.name) > 31:
            sheet = self.wb.create_sheet(title=self.name[:30])
        else:
            sheet = self.wb.create_sheet(title=self.name)
        nodes = data['data']['nodes']
        if data['data']['nodes']:
            count = 1
            sheet.cell(row=count, column=1).value = self.name
            count += 1
            for node_id, node_data in nodes.items():
                sheet.cell(row=count, column=1).value = (node_data['label'])
                count += 1

    def parse_adcs_edges(self, data) -> dict:
        nodes = data['data']['nodes']
        edges = data['data']['edges']
        parsed_data = dict()
        for node_id, node_data in nodes.items():
            parsed_data[node_id] = dict()
            parsed_data[node_id]['label'] = node_data['label']
            parsed_data[node_id]['kind'] = node_data['kind']

        for edge in edges:
            if "PublishedTo" == edge['kind']:
                parsed_data[edge['source']]['edges'] = parsed_data[edge['source']]['label'] + " is " + edge['label'] + " " + parsed_data[edge['target']]['label']
            elif "Enroll" == edge['kind']:
                parsed_data[edge['source']]['edges'] = parsed_data[edge['source']]['label'] + " can enroll using " + parsed_data[edge['target']]['label']

        return parsed_data

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--id", help="Key ID to use for bloodhound authentication generated by bloodhound CE server.")
    parser.add_argument("-k", "--key", help="Key to use for bloodhound authentication generated by bloodhound CE server.")
    parser.add_argument("-d", "--domain", help="Server to connect to: default localhost", default="localhost")
    parser.add_argument("-s", "--scheme", help="Scheme to use for connection: default http", default="http")
    parser.add_argument("-p", "--port", help="Port to connect to: default 8080", default="8080")
    parser.add_argument("-o", "--output", help="Output file to write to: default output.xlsx", default="output.xlsx")
    args = parser.parse_args()

    try:
        wb = Workbook()
    except FileNotFoundError:
        print(f"Could not open file {args.output} for writing")
        return

    # This might be best loaded from a file
    credentials = Credentials(
        token_id=args.id,
        token_key=args.key,
    )

    # Create the client and perform an example call using token request signing
    client = Client(scheme=args.scheme, host=args.domain, port=args.port, credentials=credentials)

    cypher_file = open("cyphers.json", "r")
    content = cypher_file.read()
    cyphers = json.loads(content)['cyphers']
    for cypher in cyphers:
        parser = Parser(cypher['name'], wb, client, args.output)
        print(f"[*] Running Cypher: {cypher['name']}")
        cypher_result = client.run_cypher(cypher['cypher'])
        parser.parse(cypher_result)

    wb.save(args.output)


if __name__ == "__main__":
    main()
